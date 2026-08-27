import json
from pathlib import Path

from openpyxl import load_workbook


BASE_DIR = Path(__file__).resolve().parent.parent

EXCEL_FILE = BASE_DIR / "AI_QE_Lab_Datasets_and_Governance.xlsx"
OUTPUT_FILE = BASE_DIR / "datasets" / "evaluation_dataset.json"
SHEET_NAME = "Evaluation Dataset"


def normalize_value(value):
    if value is None:
        return None

    if isinstance(value, str):
        return value.strip()

    return value


def export_dataset():
    workbook = load_workbook(
        EXCEL_FILE,
        data_only=True,
    )

    if SHEET_NAME not in workbook.sheetnames:
        raise ValueError(
            f"Sheet '{SHEET_NAME}' not found. "
            f"Available sheets: {workbook.sheetnames}"
        )

    sheet = workbook[SHEET_NAME]

    rows = list(sheet.iter_rows(values_only=True))

    if not rows:
        raise ValueError("Evaluation Dataset sheet is empty")

    headers = [
        str(header).strip()
        if header is not None
        else ""
        for header in rows[0]
    ]

    cases = []

    for row in rows[1:]:
        if not any(value is not None for value in row):
            continue

        case = {}

        for header, value in zip(headers, row):
            if not header:
                continue

            case[header] = normalize_value(value)

        cases.append(case)

    OUTPUT_FILE.parent.mkdir(
        parents=True,
        exist_ok=True,
    )

    with open(
        OUTPUT_FILE,
        "w",
        encoding="utf-8",
    ) as file:
        json.dump(
            cases,
            file,
            ensure_ascii=False,
            indent=2,
        )

    print(f"Exported cases: {len(cases)}")
    print(f"Output: {OUTPUT_FILE}")


if __name__ == "__main__":
    export_dataset()